from __future__ import annotations

from dataclasses import dataclass

import openpyxl


@dataclass
class WorkbookLocation:
    sheet: str
    cell: str


def _resolve_named_range(wb: openpyxl.Workbook, name: str) -> WorkbookLocation:
    dn = wb.defined_names.get(name)
    if dn is None:
        raise KeyError(f"NamedRange not found: {name}")
    dests = list(dn.destinations)
    if len(dests) != 1:
        raise ValueError(
            f"NamedRange '{name}' resolves to {len(dests)} destinations; only single-cell ranges are supported"
        )
    sheet, cell = dests[0]
    return WorkbookLocation(sheet=sheet, cell=cell)


def get_value(
    wb_path: str,
    mapping_type: str,
    sheet: str | None,
    cell: str | None,
    named_range: str | None,
    data_only: bool = True,
):
    wb = openpyxl.load_workbook(wb_path, data_only=data_only)
    try:
        if mapping_type == "NamedRange":
            if not named_range:
                raise ValueError("NamedRange mapping requires named_range")
            loc = _resolve_named_range(wb, named_range)
            ws = wb[loc.sheet]
            return ws[loc.cell].value
        if mapping_type == "Cell":
            if not sheet or not cell:
                raise ValueError("Cell mapping requires sheet and cell")
            ws = wb[sheet]
            return ws[cell].value
        raise ValueError(f"Unknown MappingType: {mapping_type}")
    finally:
        wb.close()


def set_value(
    wb: openpyxl.Workbook,
    mapping_type: str,
    sheet: str | None,
    cell: str | None,
    named_range: str | None,
    new_value,
):
    if mapping_type == "NamedRange":
        if not named_range:
            raise ValueError("NamedRange mapping requires named_range")
        loc = _resolve_named_range(wb, named_range)
        ws = wb[loc.sheet]
        ws[loc.cell].value = new_value
        return (loc.sheet, loc.cell)
    if mapping_type == "Cell":
        if not sheet or not cell:
            raise ValueError("Cell mapping requires sheet and cell")
        ws = wb[sheet]
        ws[cell].value = new_value
        return (sheet, cell)
    raise ValueError(f"Unknown MappingType: {mapping_type}")
