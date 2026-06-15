#!/usr/bin/env python3
"""Apply driver updates to an Excel operating model via a driver registry.

This script is intentionally limited to **explicit** mapped inputs:
- NamedRange (single-cell) OR
- Worksheet + Cell

It does NOT attempt to recalc formulas (Excel should recalc on open).

Outputs:
- Updated model copy with version tag in filename
- output/audit/ChangeLog.csv (and optional ChangeLog worksheet)
"""

from __future__ import annotations

import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

if __name__ == "__main__" and any(arg in {"-h", "--help"} for arg in sys.argv[1:]):
    print("Usage: python scripts/apply_model_updates.py plan.json")
    print("Apply explicit driver updates to an Excel model via a driver registry.")
    raise SystemExit(0)

import openpyxl
import pandas as pd

try:
    from .utils.excel_utils import set_value
    from .utils.io_utils import ensure_dir, read_json, write_text
    from .utils.validation_utils import as_float_or_none, is_missing
except ImportError:
    from utils.excel_utils import set_value
    from utils.io_utils import ensure_dir, read_json
    from utils.validation_utils import as_float_or_none, is_missing


def _read_csv(path: str) -> pd.DataFrame:
    return pd.read_csv(path, dtype=str, keep_default_na=False)


def _now_iso(tz: str | None = None) -> str:
    # Use local time; timezone label is informational.
    return datetime.now().isoformat(timespec="seconds")


def _make_version_tag(plan: dict[str, Any]) -> str:
    ev = plan.get("event", {})
    ticker = ev.get("ticker", "MISSING")
    period = ev.get("fiscal_period", "MISSING")
    tz = (
        plan.get("outputs", {})
        .get("model_update", {})
        .get("version_tag_timezone", ev.get("timezone", ""))
    )
    ts = datetime.now().strftime("%Y.%m.%d_%H%M")
    tz_lbl = tz.split("/")[-1] if tz else "LOCAL"
    return f"{ticker}_Model_v{ts}{tz_lbl}_{period}_PostEarnings"


def _safe_name(x: str) -> str:
    return "".join([c for c in x if c.isalnum() or c in "._-"])


def _append_changelog_sheet(wb: openpyxl.Workbook, changelog_df: pd.DataFrame) -> None:
    ws_name = "ChangeLog"
    if ws_name in wb.sheetnames:
        ws = wb[ws_name]
    else:
        ws = wb.create_sheet(ws_name)
        ws.append(list(changelog_df.columns))

    # Find first empty row
    row_idx = ws.max_row + 1
    for _, row in changelog_df.iterrows():
        ws.append([row.get(c) for c in changelog_df.columns])


def apply_model_updates(plan: dict[str, Any]) -> tuple[Path, Path]:
    out_dir = Path(plan.get("outputs", {}).get("output_dir", "output"))
    ensure_dir(str(out_dir / "model"))
    ensure_dir(str(out_dir / "audit"))

    model = plan.get("inputs", {}).get("model", {})
    prior_model = model.get("prior_model_xlsx")
    driver_reg_path = model.get("driver_registry_csv")

    norm = plan.get("inputs", {}).get("normalized", {})
    driver_updates_path = norm.get("driver_updates_csv")

    if is_missing(prior_model) or is_missing(driver_reg_path) or is_missing(driver_updates_path):
        raise ValueError(
            "Model update requires prior_model_xlsx, driver_registry_csv, and driver_updates_csv"
        )

    prior_model_path = Path(prior_model)
    if not prior_model_path.exists():
        raise FileNotFoundError(f"Prior model not found: {prior_model}")

    version_tag = _make_version_tag(plan)
    ext = prior_model_path.suffix
    new_name = _safe_name(version_tag) + ext
    new_model_path = out_dir / "model" / new_name

    dry_run = bool(plan.get("controls", {}).get("dry_run", False))
    if not dry_run:
        shutil.copy2(prior_model_path, new_model_path)

    driver_reg = _read_csv(driver_reg_path)
    updates = _read_csv(driver_updates_path)

    # Index registry by DriverID
    reg = {str(r["DriverID"]): r for _, r in driver_reg.iterrows()}

    changelog_rows = []
    changed_by = plan.get("controls", {}).get("prepared_by", "AI")
    model_version_id = version_tag

    keep_vba = ext.lower() in {".xlsm", ".xltm"}
    wb = openpyxl.load_workbook(
        new_model_path if not dry_run else prior_model_path, keep_vba=keep_vba
    )

    try:
        for _, u in updates.iterrows():
            did = str(u.get("DriverID") or "").strip()
            if not did:
                continue
            if did not in reg:
                # Unknown driver: skip but record
                changelog_rows.append(
                    {
                        "Timestamp": _now_iso(),
                        "ModelVersionID": model_version_id,
                        "ChangedBy": changed_by,
                        "Section": "Drivers",
                        "ItemChanged": did,
                        "OldValue": "MISSING",
                        "NewValue": u.get("NewValue"),
                        "Why": u.get("Why"),
                        "SourceTag": u.get("SourceTag"),
                    }
                )
                continue

            new_val_raw = u.get("NewValue")
            if is_missing(new_val_raw):
                continue

            r = reg[did]
            mapping_type = str(r.get("MappingType") or "").strip() or "NamedRange"
            sheet = str(r.get("Worksheet") or "").strip() or None
            cell = str(r.get("Cell") or "").strip() or None
            named_range = str(r.get("NamedRange") or "").strip() or None

            # Read old value (from mapped cell)
            try:
                if mapping_type == "NamedRange" and named_range:
                    # resolve and read via openpyxl
                    dn = wb.defined_names.get(named_range)
                    old_val = None
                    if dn:
                        dests = list(dn.destinations)
                        if len(dests) == 1:
                            sname, addr = dests[0]
                            old_val = wb[sname][addr].value
                elif mapping_type == "Cell" and sheet and cell:
                    old_val = wb[sheet][cell].value
                else:
                    old_val = None
            except Exception:
                old_val = None

            # Parse numeric if possible
            new_val_num = as_float_or_none(new_val_raw)
            new_val = new_val_num if new_val_num is not None else new_val_raw

            # Write
            try:
                ws_name, ws_cell = set_value(wb, mapping_type, sheet, cell, named_range, new_val)
            except Exception as e:
                changelog_rows.append(
                    {
                        "Timestamp": _now_iso(),
                        "ModelVersionID": model_version_id,
                        "ChangedBy": changed_by,
                        "Section": "Drivers",
                        "ItemChanged": did,
                        "OldValue": str(old_val) if old_val is not None else "MISSING",
                        "NewValue": str(new_val_raw),
                        "Why": u.get("Why"),
                        "SourceTag": u.get("SourceTag"),
                    }
                )
                continue

            changelog_rows.append(
                {
                    "Timestamp": _now_iso(),
                    "ModelVersionID": model_version_id,
                    "ChangedBy": changed_by,
                    "Section": "Drivers",
                    "ItemChanged": did,
                    "OldValue": str(old_val) if old_val is not None else "MISSING",
                    "NewValue": str(new_val_raw),
                    "Why": u.get("Why"),
                    "SourceTag": u.get("SourceTag"),
                }
            )

        changelog_df = pd.DataFrame(
            changelog_rows,
            columns=[
                "Timestamp",
                "ModelVersionID",
                "ChangedBy",
                "Section",
                "ItemChanged",
                "OldValue",
                "NewValue",
                "Why",
                "SourceTag",
            ],
        )

        changelog_path = out_dir / "audit" / "ChangeLog.csv"
        changelog_df.to_csv(changelog_path, index=False)

        # Optional: add/append ChangeLog tab
        try:
            if len(changelog_df) > 0:
                _append_changelog_sheet(wb, changelog_df)
        except Exception:
            pass

        if not dry_run:
            wb.save(new_model_path)

        return new_model_path, changelog_path
    finally:
        wb.close()


def main() -> int:
    if len(sys.argv) != 2:
        print("Usage: python scripts/apply_model_updates.py plan.json")
        return 1

    plan = read_json(sys.argv[1])
    new_model, changelog = apply_model_updates(plan)
    print(f"Wrote updated model: {new_model}")
    print(f"Wrote changelog: {changelog}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
