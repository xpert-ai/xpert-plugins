#!/usr/bin/env python3
"""
Static workbook inspection helper for the model-audit-tieout skill.

This script does not calculate Excel formulas. It inventories workbook structure,
formulas, visible/hidden sheets, hardcodes inside formulas, external links,
volatile functions, and simple formula-family inconsistencies. Use the output as
an audit starting point, then apply financial-modeling judgment manually.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, Iterable

PLUGIN_ROOT = Path(__file__).resolve().parents[3]
if str(PLUGIN_ROOT) not in sys.path:
    sys.path.insert(0, str(PLUGIN_ROOT))

from shared.artifact_packager import (  # noqa: E402
    artifact_item,
    dict_rows_to_sheet,
    logs_dir,
    support_dir,
    write_artifact_manifest,
    write_cover_first_workbook,
    write_dashboard_contract,
    write_report_html,
)

VOLATILE_FUNCTIONS = [
    "TODAY",
    "NOW",
    "RAND",
    "RANDBETWEEN",
    "OFFSET",
    "INDIRECT",
    "CELL",
    "INFO",
]

ERROR_LITERALS = ["#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!"]
COMMON_CONSTANTS = {0, 1, -1, 2, -2, 10, 12, 52, 100, 365, 360, 1000, 1000000}

CELL_REF_RE = re.compile(r"(?<![A-Z0-9_])\$?[A-Z]{1,3}\$?\d+(?![A-Z0-9_])")
RANGE_RE = re.compile(r"\$?[A-Z]{1,3}\$?\d+\s*:\s*\$?[A-Z]{1,3}\$?\d+", re.IGNORECASE)
SHEET_REF_RE = re.compile(r"(?:'[^']+'|[A-Za-z0-9_ ]+)!\$?[A-Z]{1,3}\$?\d+", re.IGNORECASE)
EXTERNAL_REF_RE = re.compile(
    r"\[[^\]]+\]|\.xlsx|\.xlsm|\.xls|http[s]?://|\\\\|/[A-Za-z0-9_. -]+/", re.IGNORECASE
)
NUMERIC_LITERAL_RE = re.compile(r"(?<![A-Za-z0-9_\$])[-+]?\d+(?:\.\d+)?%?(?![A-Za-z0-9_])")
DATE_HINT_RE = re.compile(
    r"\b(as of|updated|last updated|date|period|source date|market data)\b", re.IGNORECASE
)


@dataclass
class WorkbookMapRow:
    sheet: str
    state: str
    max_row: int
    max_column: int
    used_cells: int
    formula_cells: int
    constant_cells: int
    blank_cells_in_used_range: int
    merged_ranges: int
    freeze_panes: str
    has_tables: bool
    has_auto_filter: bool


@dataclass
class FormulaRow:
    sheet: str
    cell: str
    formula: str
    normalized_formula: str
    formula_length: int
    cached_value: str
    has_external_ref: bool
    volatile_functions: str
    has_error_literal: bool
    hardcoded_numbers: str


@dataclass
class IssueRow:
    severity: str
    category: str
    sheet: str
    cell_or_range: str
    finding: str
    why_it_matters: str
    recommended_next_step: str


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def strip_strings(formula: str) -> str:
    # Replace double-quoted strings to avoid false positives in formulas.
    return re.sub(r'"(?:[^"]|"")*"', '"STR"', formula)


def normalize_formula(formula: str) -> str:
    s = strip_strings(formula.upper())
    s = RANGE_RE.sub("RANGE", s)
    s = SHEET_REF_RE.sub("SHEET!REF", s)
    s = CELL_REF_RE.sub("REF", s)
    s = re.sub(r"\s+", "", s)
    return s


def volatile_functions(formula: str) -> list[str]:
    found = []
    upper = formula.upper()
    for fn in VOLATILE_FUNCTIONS:
        if re.search(r"\b" + re.escape(fn) + r"\s*\(", upper):
            found.append(fn)
    return found


def extract_hardcoded_numbers(formula: str) -> list[str]:
    s = strip_strings(formula.upper())
    # Remove cell refs/ranges first to avoid row numbers appearing as numeric literals.
    s = RANGE_RE.sub(" RANGE ", s)
    s = SHEET_REF_RE.sub(" SHEET_REF ", s)
    s = CELL_REF_RE.sub(" REF ", s)
    raw = NUMERIC_LITERAL_RE.findall(s)
    values = []
    for item in raw:
        cleaned = item.rstrip("%")
        try:
            num = float(cleaned)
        except ValueError:
            continue
        if num.is_integer():
            comp = int(num)
        else:
            comp = num
        if comp in COMMON_CONSTANTS:
            continue
        # Keep percentages and non-common numbers; these are often assumptions.
        values.append(item)
    # Deduplicate while preserving order.
    seen = set()
    result = []
    for val in values:
        if val not in seen:
            seen.add(val)
            result.append(val)
    return result


def write_csv(path: Path, rows: Iterable[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def add_issue(
    issues: list[IssueRow],
    severity: str,
    category: str,
    sheet: str,
    cell_or_range: str,
    finding: str,
    why: str,
    next_step: str,
) -> None:
    issues.append(
        IssueRow(
            severity=severity,
            category=category,
            sheet=sheet,
            cell_or_range=cell_or_range,
            finding=finding,
            why_it_matters=why,
            recommended_next_step=next_step,
        )
    )


def scan_workbook(
    workbook_path: Path,
) -> tuple[list[WorkbookMapRow], list[FormulaRow], list[IssueRow], dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        print("ERROR: openpyxl is required to run this script.", file=sys.stderr)
        print("Install with: python -m pip install -r scripts/requirements.txt", file=sys.stderr)
        print(str(exc), file=sys.stderr)
        sys.exit(2)

    wb_formula = load_workbook(workbook_path, data_only=False, keep_links=True)
    try:
        wb_values = load_workbook(workbook_path, data_only=True, keep_links=True)
    except Exception:
        wb_values = None

    workbook_map: list[WorkbookMapRow] = []
    formulas: list[FormulaRow] = []
    issues: list[IssueRow] = []
    date_hints: list[dict[str, str]] = []
    formula_cells_by_sheet_row: dict[tuple[str, int], list[tuple[str, str, str]]] = defaultdict(
        list
    )

    external_links_count = len(getattr(wb_formula, "_external_links", []) or [])
    if external_links_count:
        add_issue(
            issues,
            "high",
            "formula_integrity",
            "workbook",
            "external links",
            f"workbook contains {external_links_count} external link object(s)",
            "external links can make audit results stale or unreproducible if source files are missing",
            "review linked files, refresh status, and tie material linked values to uploaded sources",
        )

    if getattr(wb_formula.calculation, "iterate", False):
        add_issue(
            issues,
            "high",
            "model_architecture",
            "workbook",
            "calculation settings",
            "iterative calculation appears enabled",
            "iterative settings can hide circularity in interest, debt, cash sweep, or plug logic",
            "document intentional circular references and test model behavior with iteration settings confirmed",
        )

    visible_sheets = [ws for ws in wb_formula.worksheets if ws.sheet_state == "visible"]
    first_visible = visible_sheets[0] if visible_sheets else None
    normalized_cover_title = (
        first_visible.title.lower().replace(" ", "_") if first_visible is not None else ""
    )
    is_update_cover = re.fullmatch(r"update_cover(?:_\d+)?", normalized_cover_title) is not None
    if first_visible is None:
        add_issue(
            issues,
            "high",
            "presentation_readiness",
            "workbook",
            "visible sheets",
            "workbook has no visible sheets",
            "reviewers cannot inspect key outputs, sources, or checks without changing workbook state",
            "restore a visible review surface before relying on the workbook",
        )
    elif normalized_cover_title != "cover" and not is_update_cover:
        add_issue(
            issues,
            "medium",
            "presentation_readiness",
            first_visible.title,
            "first visible sheet",
            "first visible sheet is not a Cover tab",
            "PM and senior-review workbooks should open to a Cover insight dashboard before raw model, README, or control tabs",
            "add a first-tab Cover tied to model outputs, sources, checks, and the investment decision",
        )
    else:
        cover_text = " ".join(
            stringify(value).lower()
            for row in first_visible.iter_rows(max_row=60, max_col=10, values_only=True)
            for value in row
            if value not in (None, "")
        )
        required_terms = {
            "executive read-through": [
                "executive read-through",
                "net read",
                "decision question",
                "recommendation",
            ],
            "headline output": [
                "headline output",
                "selected valuation",
                "value per share",
                "base revenue",
                "liquidity trough",
                "thesis status",
            ],
            "key metrics": [
                "kpi",
                "key metric",
                "metric",
                "revenue",
                "ebitda",
                "fcf",
                "cash",
                "multiple",
            ],
            "scenario or output table": [
                "scenario",
                "base",
                "upside",
                "downside",
                "bull",
                "bear",
                "output",
            ],
            "source posture": ["source posture", "source count", "source", "evidence"],
            "warnings or failures": ["warning", "hard failure", "qa", "check"],
            "chart-ready or visual data": ["chart-ready", "chart ready", "visual", "sensitivity"],
            "workbook map": ["workbook map", "tab map"],
        }
        if is_update_cover:
            required_terms["executive read-through"].extend(["bottom line", "pm posture"])
            required_terms["headline output"].extend(["estimate change", "valuation change"])
            required_terms["scenario or output table"].extend(["rebuild", "blocked"])
            required_terms["warnings or failures"].extend(["blocked", "stale"])
            required_terms["chart-ready or visual data"].extend(["revenue period", "latest value"])
        missing = [
            label
            for label, terms in required_terms.items()
            if not any(term in cover_text for term in terms)
        ]
        if missing:
            add_issue(
                issues,
                "medium",
                "presentation_readiness",
                first_visible.title,
                "cover contents",
                f"cover/dashboard may be missing: {', '.join(missing)}",
                "a Cover can exist mechanically but still fail as an insight dashboard if the investment read-through, metrics, sources, warnings, visuals, or navigation are buried",
                "expand the first tab to include net read, headline outputs, key metrics, scenarios, source posture, warnings/hard failures, chart-ready data, and workbook map",
            )

    for ws in wb_formula.worksheets:
        ws_values = wb_values[ws.title] if wb_values and ws.title in wb_values.sheetnames else None
        used_cells = 0
        formula_cells = 0
        constant_cells = 0
        blank_cells = 0

        if ws.sheet_state != "visible":
            sev = "high" if ws.sheet_state == "veryHidden" else "medium"
            add_issue(
                issues,
                sev,
                "model_architecture",
                ws.title,
                "sheet visibility",
                f"sheet is {ws.sheet_state}",
                "hidden sheets can contain source data, assumptions, checks, or formulas not visible to reviewers",
                "inspect hidden sheet contents and disclose whether it feeds material outputs",
            )

        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                used_cells += 1
                coord = cell.coordinate
                if is_formula(val):
                    formula_cells += 1
                    cached = ""
                    if ws_values is not None:
                        try:
                            cached = stringify(ws_values[coord].value)
                        except Exception:
                            cached = ""
                    norm = normalize_formula(val)
                    vols = volatile_functions(val)
                    hardcodes = extract_hardcoded_numbers(val)
                    has_external = bool(EXTERNAL_REF_RE.search(val))
                    has_error = any(err in val.upper() for err in ERROR_LITERALS)
                    formulas.append(
                        FormulaRow(
                            sheet=ws.title,
                            cell=coord,
                            formula=val,
                            normalized_formula=norm,
                            formula_length=len(val),
                            cached_value=cached,
                            has_external_ref=has_external,
                            volatile_functions=";".join(vols),
                            has_error_literal=has_error,
                            hardcoded_numbers=";".join(hardcodes),
                        )
                    )
                    formula_cells_by_sheet_row[(ws.title, cell.row)].append((coord, val, norm))

                    if has_external:
                        add_issue(
                            issues,
                            "high",
                            "formula_integrity",
                            ws.title,
                            coord,
                            "formula contains an external reference",
                            "external references can produce stale or unreproducible values if the linked file is unavailable or outdated",
                            "tie the value to the linked source file or replace with documented source-tab data",
                        )
                    if vols:
                        add_issue(
                            issues,
                            "medium",
                            "formula_integrity",
                            ws.title,
                            coord,
                            f"formula uses volatile function(s): {', '.join(vols)}",
                            "volatile functions can change outputs on recalculation and complicate audit reproducibility",
                            "confirm volatility is intentional and does not drive material outputs without disclosure",
                        )
                    if hardcodes:
                        sev = "medium" if len(hardcodes) <= 2 else "high"
                        add_issue(
                            issues,
                            sev,
                            "formula_integrity",
                            ws.title,
                            coord,
                            f"formula contains review-worthy numeric literal(s): {', '.join(hardcodes[:8])}",
                            "hardcoded numbers inside formulas may hide assumptions, rates, multiples, or adjustments",
                            "move material assumptions to an input/source tab or document why the constant is appropriate",
                        )
                    if has_error:
                        add_issue(
                            issues,
                            "high",
                            "formula_integrity",
                            ws.title,
                            coord,
                            "formula contains an error literal",
                            "error literals may indicate broken references or intentionally suppressed errors",
                            "inspect precedent cells and resolve or document the error treatment",
                        )
                else:
                    constant_cells += 1
                    if isinstance(val, str) and DATE_HINT_RE.search(val):
                        # Capture local context around date/source labels for manual review.
                        date_hints.append({"sheet": ws.title, "cell": coord, "text": val[:200]})
                # blanks in used range are computed below

        total_range_cells = ws.max_row * ws.max_column if ws.max_row and ws.max_column else 0
        blank_cells = max(total_range_cells - used_cells, 0)
        workbook_map.append(
            WorkbookMapRow(
                sheet=ws.title,
                state=ws.sheet_state,
                max_row=ws.max_row,
                max_column=ws.max_column,
                used_cells=used_cells,
                formula_cells=formula_cells,
                constant_cells=constant_cells,
                blank_cells_in_used_range=blank_cells,
                merged_ranges=len(list(ws.merged_cells.ranges)),
                freeze_panes=stringify(ws.freeze_panes),
                has_tables=bool(getattr(ws, "tables", None)),
                has_auto_filter=bool(ws.auto_filter and ws.auto_filter.ref),
            )
        )

        if len(list(ws.merged_cells.ranges)) > 0:
            add_issue(
                issues,
                "low",
                "formatting",
                ws.title,
                "merged ranges",
                f"sheet contains {len(list(ws.merged_cells.ranges))} merged range(s)",
                "merged cells can complicate formula review, data extraction, and tie-out automation",
                "review whether merged ranges are cosmetic or interfere with source data / calculations",
            )

    # Identify simple formula-family inconsistencies by row.
    for (sheet, rownum), row_formulas in formula_cells_by_sheet_row.items():
        if len(row_formulas) < 4:
            continue
        norms = [item[2] for item in row_formulas]
        counts = Counter(norms)
        dominant_norm, dominant_count = counts.most_common(1)[0]
        # Flag minority formulas when a clear dominant pattern exists.
        if dominant_count >= max(3, int(len(row_formulas) * 0.65)) and len(counts) > 1:
            minority = [
                (coord, formula) for coord, formula, norm in row_formulas if norm != dominant_norm
            ]
            if 0 < len(minority) <= 5:
                cells = ", ".join(coord for coord, _ in minority)
                add_issue(
                    issues,
                    "medium",
                    "formula_integrity",
                    sheet,
                    f"row {rownum}: {cells}",
                    "row contains formula(s) that differ from the dominant copied formula pattern",
                    "unexpected formula-family breaks can indicate overwritten formulas, wrong period references, or hidden assumptions",
                    "inspect the minority cells and confirm whether the break is intentional",
                )

    # Flag sheets that are mostly constants but named like assumptions/source checks.
    for row in workbook_map:
        name = row.sheet.lower()
        if (
            row.used_cells
            and row.formula_cells == 0
            and any(token in name for token in ["source", "input", "assumption", "data"])
        ):
            add_issue(
                issues,
                "info",
                "source_tieout",
                row.sheet,
                "entire sheet",
                "sheet appears to be a source/input tab with no formulas",
                "source/input tabs should be tied to underlying documents and as-of dates",
                "verify source labels, source dates, and tie-outs for material values on this tab",
            )

    summary = {
        "workbook": str(workbook_path.name),
        "sheets": len(wb_formula.worksheets),
        "hidden_sheets": sum(1 for r in workbook_map if r.state != "visible"),
        "formula_cells": len(formulas),
        "external_link_objects": external_links_count,
        "issues_by_severity": dict(Counter(issue.severity for issue in issues)),
        "issues_by_category": dict(Counter(issue.category for issue in issues)),
        "date_source_hints": date_hints[:100],
        "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "notes": [
            "static inspection only; this script does not calculate formulas or prove business logic is correct",
            "use outputs as a triage starting point and supplement with manual model review",
        ],
    }
    return workbook_map, formulas, issues, summary


def make_markdown_report(
    workbook_path: Path,
    workbook_map: list[WorkbookMapRow],
    formulas: list[FormulaRow],
    issues: list[IssueRow],
    summary: dict[str, Any],
) -> str:
    sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
    sorted_issues = sorted(
        issues, key=lambda x: (sev_order.get(x.severity, 99), x.category, x.sheet, x.cell_or_range)
    )
    visible_issue_count = len(sorted_issues)
    high_count = sum(1 for issue in sorted_issues if issue.severity in {"critical", "high"})
    posture = (
        "red"
        if any(i.severity == "critical" for i in sorted_issues)
        else "yellow"
        if high_count
        else "green"
    )
    if not workbook_map:
        posture = "gray"

    lines: list[str] = []
    lines.append(f"# Workbook Audit Report: {workbook_path.name}")
    lines.append("")
    lines.append("## Static inspection summary")
    lines.append("")
    lines.append(f"- **Generated:** {summary.get('generated_at', '')}")
    lines.append(f"- **Sheets:** {summary.get('sheets', 0)}")
    lines.append(f"- **Hidden sheets:** {summary.get('hidden_sheets', 0)}")
    lines.append(f"- **Formula cells:** {summary.get('formula_cells', 0)}")
    lines.append(f"- **External link objects:** {summary.get('external_link_objects', 0)}")
    lines.append(f"- **Initial model health posture:** {posture}")
    lines.append("")
    lines.append(
        "> This report is a static inspection only. It does not calculate Excel formulas or validate investment logic. Use it as an audit starting point."
    )
    lines.append("")

    lines.append("## Workbook map")
    lines.append("")
    lines.append(
        "| sheet | state | rows | columns | used cells | formulas | constants | merged ranges |"
    )
    lines.append("|---|---|---:|---:|---:|---:|---:|---:|")
    for row in workbook_map:
        lines.append(
            f"| {row.sheet} | {row.state} | {row.max_row} | {row.max_column} | {row.used_cells} | {row.formula_cells} | {row.constant_cells} | {row.merged_ranges} |"
        )
    lines.append("")

    lines.append("## Priority issue log")
    lines.append("")
    if visible_issue_count == 0:
        lines.append(
            "No mechanical issues were flagged by static inspection. Manual model audit is still required for business logic, source tie-outs, and decision-readiness."
        )
    else:
        lines.append("| severity | category | location | finding | recommended next step |")
        lines.append("|---|---|---|---|---|")
        for issue in sorted_issues[:100]:
            location = f"{issue.sheet} {issue.cell_or_range}".strip()
            lines.append(
                f"| {issue.severity} | {issue.category} | {location} | {issue.finding} | {issue.recommended_next_step} |"
            )
        if visible_issue_count > 100:
            lines.append(
                f"| info | summary | issue log | {visible_issue_count - 100} additional issues omitted from audit report | see issues.csv |"
            )
    lines.append("")

    top_hardcodes = [f for f in formulas if f.hardcoded_numbers]
    top_external = [f for f in formulas if f.has_external_ref]
    top_volatile = [f for f in formulas if f.volatile_functions]

    lines.append("## Formula exception summary")
    lines.append("")
    lines.append(f"- formulas with review-worthy hardcoded numeric literals: {len(top_hardcodes)}")
    lines.append(f"- formulas with external references: {len(top_external)}")
    lines.append(f"- formulas with volatile functions: {len(top_volatile)}")
    lines.append("")
    lines.append("See formulas.csv for full formula inventory.")
    lines.append("")

    if summary.get("date_source_hints"):
        lines.append("## Date/source label hints")
        lines.append("")
        lines.append(
            "These cells may contain source-date or as-of-date labels worth reviewing during source tie-out."
        )
        lines.append("")
        lines.append("| sheet | cell | text |")
        lines.append("|---|---|---|")
        for hint in summary["date_source_hints"][:25]:
            text = hint.get("text", "").replace("|", " ")
            lines.append(f"| {hint.get('sheet', '')} | {hint.get('cell', '')} | {text} |")
        lines.append("")

    lines.append("## Recommended next audit steps")
    lines.append("")
    lines.append("1. Trace material outputs to precedent formulas and source tabs.")
    lines.append("2. Build a source tie-out ledger for all decision-driving values.")
    lines.append(
        "3. Review assumptions and sensitivities for the model's actual public-equity investment decision."
    )
    lines.append(
        "4. Resolve critical/high issues before using the model in an PM, IC, client, or trading decision."
    )
    lines.append("")
    return "\n".join(lines)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Static audit of an Excel workbook for model-audit-tieout."
    )
    parser.add_argument("workbook", help="Path to .xlsx or .xlsm workbook")
    parser.add_argument(
        "--out-dir", default="audit_output", help="Directory where audit outputs will be written"
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.exists():
        print(f"ERROR: workbook not found: {workbook_path}", file=sys.stderr)
        return 1
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        print(
            "ERROR: only .xlsx and .xlsm files are supported by this static inspector.",
            file=sys.stderr,
        )
        return 1

    out_dir = Path(args.out_dir).expanduser().resolve()
    workbook_map, formulas, issues, summary = scan_workbook(workbook_path)

    out_dir.mkdir(parents=True, exist_ok=True)
    csv_dir = support_dir(out_dir)
    log_dir = logs_dir(out_dir)
    workbook_map_rows = [asdict(row) for row in workbook_map]
    formula_rows = [asdict(row) for row in formulas]
    issue_rows = [asdict(row) for row in issues]

    write_csv(
        csv_dir / "workbook_map.csv",
        workbook_map_rows,
        list(asdict(workbook_map[0]).keys())
        if workbook_map
        else [
            "sheet",
            "state",
            "max_row",
            "max_column",
            "used_cells",
            "formula_cells",
            "constant_cells",
            "blank_cells_in_used_range",
            "merged_ranges",
            "freeze_panes",
            "has_tables",
            "has_auto_filter",
        ],
    )
    write_csv(
        csv_dir / "formula_exception_log.csv",
        formula_rows,
        list(asdict(formulas[0]).keys())
        if formulas
        else [
            "sheet",
            "cell",
            "formula",
            "normalized_formula",
            "formula_length",
            "cached_value",
            "has_external_ref",
            "volatile_functions",
            "has_error_literal",
            "hardcoded_numbers",
        ],
    )
    write_csv(
        csv_dir / "model_audit_findings.csv",
        issue_rows,
        list(asdict(issues[0]).keys())
        if issues
        else [
            "severity",
            "category",
            "sheet",
            "cell_or_range",
            "finding",
            "why_it_matters",
            "recommended_next_step",
        ],
    )
    source_tieout_rows = [
        {
            "sheet": hint.get("sheet", ""),
            "cell_or_range": hint.get("cell", ""),
            "source_or_date_label": hint.get("text", ""),
            "tie_out_status": "needs_review",
            "review_note": "Source/date label detected by static scan; tie to controlling filing, market data, model source, or assumption support.",
        }
        for hint in summary.get("date_source_hints", [])
    ]
    write_csv(
        csv_dir / "source_tieout_ledger.csv",
        source_tieout_rows,
        ["sheet", "cell_or_range", "source_or_date_label", "tie_out_status", "review_note"],
    )
    with (log_dir / "model_audit_findings.json").open("w", encoding="utf-8") as f:
        json.dump(
            {
                "summary": summary,
                "issues": issue_rows,
                "formula_exceptions": formula_rows,
                "source_tieout_ledger": source_tieout_rows,
            },
            f,
            indent=2,
        )

    report = make_markdown_report(workbook_path, workbook_map, formulas, issues, summary)
    support_note_path = csv_dir / "model_audit_support_note.md"
    support_note_path.write_text(report, encoding="utf-8")
    report_path = write_report_html(
        out_dir / "model_audit_report.html", "Public Equity Investing Equity Model Audit", report
    )
    issue_workbook = out_dir / "model_audit_issues.xlsx"
    write_cover_first_workbook(
        issue_workbook,
        [
            ["Model Audit Issue Log"],
            ["Audited workbook", str(workbook_path)],
            ["Sheets", summary.get("sheets", 0)],
            ["Formula cells", summary.get("formula_cells", 0)],
            ["Issues", len(issues)],
            [
                "First read",
                "Open model_audit_report.html first; this workbook is the companion issue log.",
            ],
        ],
        {
            "Issues": dict_rows_to_sheet(issue_rows),
            "Workbook_Map": dict_rows_to_sheet(workbook_map_rows),
            "Formula_Log": dict_rows_to_sheet(formula_rows),
            "Source_Tieout": dict_rows_to_sheet(source_tieout_rows),
        },
    )
    contract_path = log_dir / "model_audit_dashboard_contract.json"
    write_dashboard_contract(
        contract_path,
        "model-audit-tieout",
        "Public Equity Investing Equity Model Audit",
        workbook_path.name,
        "report_only",
        report_path,
        executive_summary="Model audit report with formula exceptions, workbook map, source tie-out hints, and prioritized fixes.",
        hero_actions=[
            {"label": "Open audit report", "path": str(report_path)},
            {"label": "Open issue workbook", "path": str(issue_workbook)},
        ],
        supporting_outputs=[
            {
                "path": str(csv_dir / "model_audit_findings.csv"),
                "label": "Issue CSV",
                "role": "support",
            }
        ],
        report_body=[{"heading": "Audit Summary", "body": report}],
        readiness_posture="needs_review" if issues else "first-pass-clear",
    )
    write_artifact_manifest(
        out_dir,
        "model-audit-tieout",
        "html_report",
        report_path,
        companion_deliverables=[
            artifact_item(
                issue_workbook,
                "companion_deliverable",
                "xlsx",
                "Cover-first issue-log workbook.",
                True,
                True,
            )
        ],
        support_artifacts=[
            artifact_item(
                csv_dir / "workbook_map.csv",
                "support_artifact",
                "csv",
                "Workbook structure inventory.",
                False,
                True,
                "CSV is audit support for filtering/import.",
            ),
            artifact_item(
                csv_dir / "formula_exception_log.csv",
                "support_artifact",
                "csv",
                "Formula inventory and exception leads.",
                False,
                True,
                "CSV backs the audit report and workbook.",
            ),
            artifact_item(
                csv_dir / "model_audit_findings.csv",
                "support_artifact",
                "csv",
                "Raw model audit issue table.",
                False,
                True,
                "CSV backs the audit report and issue workbook.",
            ),
            artifact_item(
                csv_dir / "source_tieout_ledger.csv",
                "support_artifact",
                "csv",
                "Source/date tie-out hints.",
                False,
                True,
                "CSV supports source-to-cell review.",
            ),
            artifact_item(
                csv_dir / "model_audit_support_note.md",
                "support_artifact",
                "markdown",
                "Narrative support note backing the HTML report.",
                False,
                True,
                "Markdown is narrative support, not the lead artifact.",
            ),
            artifact_item(
                log_dir / "model_audit_findings.json",
                "support_artifact",
                "json",
                "Machine-readable model audit findings.",
                False,
                True,
                "JSON is audit support for downstream tooling.",
            ),
            artifact_item(
                contract_path,
                "support_artifact",
                "json",
                "Dashboard/report contract.",
                False,
                False,
                "Contract JSON is renderer plumbing.",
            ),
        ],
        blocked_or_partial_status={
            "status": "partial",
            "reason": "Static workbook inspection does not recalculate formulas or prove investment logic.",
            "missing_inputs": [
                "Manual formula trace",
                "Source document tie-out",
                "Scenario/downside review",
            ],
        },
    )

    print(f"Wrote audit outputs to: {out_dir}")
    print(
        f"Sheets: {summary.get('sheets', 0)} | Formula cells: {summary.get('formula_cells', 0)} | Issues: {len(issues)}"
    )
    print(f"Support note: {support_note_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
