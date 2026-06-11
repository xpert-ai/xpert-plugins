#!/usr/bin/env python3
"""Create deterministic CSV/XLSX thesis tracker artifacts from structured inputs."""

from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

SECTIONS: dict[str, list[str]] = {
    "dashboard": [
        "company_issuer",
        "ticker_security",
        "direction",
        "mandate_role",
        "current_thesis_status",
        "conviction",
        "position_rating",
        "base_bull_bear_value",
        "next_catalyst",
        "action_recommendation",
        "next_review_date",
        "data_gaps",
    ],
    "thesis_pillars": [
        "pillar_id",
        "pillar_name",
        "claim",
        "weight",
        "current_status",
        "kpi_evidence",
        "confirm_threshold",
        "warning_threshold",
        "break_threshold",
        "latest_evidence",
        "signal",
        "model_linkage",
        "next_proof_point",
        "owner",
    ],
    "evidence_ledger": [
        "evidence_id",
        "date_time",
        "source",
        "source_type",
        "event",
        "evidence_fact",
        "prior_house_expectation",
        "consensus_market_expectation",
        "interpretation",
        "pillar_affected",
        "signal_direction",
        "evidence_quality",
        "model_impact",
        "valuation_impact",
        "action_implication",
        "follow_up",
    ],
    "kpi_tracker": [
        "kpi",
        "pillar",
        "source",
        "unit",
        "period",
        "baseline",
        "prior_actual",
        "current_actual",
        "house_estimate",
        "consensus",
        "threshold",
        "status",
        "trend",
        "comment",
    ],
    "catalyst_calendar": [
        "catalyst_id",
        "date_window",
        "catalyst",
        "pillar_tested",
        "expected_outcome",
        "market_setup",
        "upside_case",
        "downside_case",
        "required_prep",
        "actual_result",
        "status_action",
    ],
    "estimate_revisions": [
        "date",
        "source",
        "metric",
        "period",
        "prior_estimate",
        "new_estimate",
        "change",
        "driver",
        "thesis_implication",
    ],
    "model_changelog": [
        "version",
        "date",
        "changed_assumption",
        "old_value",
        "new_value",
        "reason",
        "impact",
        "reviewer",
        "status",
    ],
    "decision_log": [
        "date",
        "decision",
        "rationale",
        "evidence_ids",
        "thesis_status",
        "position_rating_change",
        "risk_reward",
        "pm_owner",
        "next_review",
    ],
    "sources": [
        "source_id",
        "source_name",
        "source_type",
        "date_as_of",
        "reliability",
        "used_for",
        "citation_link",
        "limitation",
    ],
    "open_questions": [
        "question_id",
        "question",
        "why_it_matters",
        "owner",
        "due_date",
        "status",
        "decision_impact",
    ],
}


ALIASES = {
    "source_inventory": "sources",
    "source_register": "sources",
    "open_items": "open_questions",
}


def normalize_key(key: str) -> str:
    return "".join(ch.lower() if ch.isalnum() else "_" for ch in key).strip("_")


def normalize_row(row: dict[str, Any], fields: list[str]) -> dict[str, Any]:
    normalized = {normalize_key(str(k)): v for k, v in row.items()}
    return {field: normalized.get(field, "") for field in fields}


def rows_for_section(
    payload: dict[str, Any], section: str, fields: list[str]
) -> list[dict[str, Any]]:
    source = payload.get(section)
    if source is None:
        for alias, target in ALIASES.items():
            if target == section and alias in payload:
                source = payload[alias]
                break
    if source is None:
        return [{field: "" for field in fields}]
    if isinstance(source, dict):
        if section == "dashboard":
            return [{"field": k, "value": v} for k, v in source.items()] or [
                {"field": "", "value": ""}
            ]
        source = [source]
    if not isinstance(source, list):
        return [{field: "" for field in fields}]
    rows = [normalize_row(item, fields) for item in source if isinstance(item, dict)]
    return rows or [{field: "" for field in fields}]


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    if rows and set(rows[0].keys()) == {"field", "value"}:
        fields = ["field", "value"]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def row_has_data(row: dict[str, Any]) -> bool:
    return any(value not in ("", None) for value in row.values())


def tracker_cover_rows(
    section_rows: dict[str, tuple[list[str], list[dict[str, Any]]]],
) -> list[dict[str, Any]]:
    dashboard_fields, dashboard_rows = section_rows.get("dashboard", ([], []))
    dashboard: dict[str, Any] = {}
    if dashboard_rows and set(dashboard_rows[0].keys()) == {"field", "value"}:
        dashboard = {
            normalize_key(str(row.get("field", ""))): row.get("value", "") for row in dashboard_rows
        }
    elif dashboard_rows:
        dashboard = dashboard_rows[0]

    def count_rows(section: str) -> int:
        rows = section_rows.get(section, ([], []))[1]
        return sum(1 for row in rows if row_has_data(row))

    return [
        {
            "section": "Header",
            "metric": "Issuer / security",
            "value": dashboard.get("company_issuer") or dashboard.get("ticker_security", ""),
            "notes": "Thesis tracker workbook landing page.",
        },
        {
            "section": "Status",
            "metric": "Current thesis status",
            "value": dashboard.get("current_thesis_status", ""),
            "notes": "Confirm whether this status is supported by the evidence ledger.",
        },
        {
            "section": "Status",
            "metric": "Conviction / position rating",
            "value": f"{dashboard.get('conviction', '')} / {dashboard.get('position_rating', '')}",
            "notes": "Keep rating, sizing, and evidence quality aligned.",
        },
        {
            "section": "Valuation frame",
            "metric": "Base / bull / bear value",
            "value": dashboard.get("base_bull_bear_value", ""),
            "notes": "Tie to model changelog or valuation artifact where available.",
        },
        {
            "section": "Catalyst dashboard",
            "metric": "Next catalyst",
            "value": dashboard.get("next_catalyst", ""),
            "notes": "Review catalyst_calendar and prep needs.",
        },
        {
            "section": "Action dashboard",
            "metric": "Action recommendation",
            "value": dashboard.get("action_recommendation", ""),
            "notes": "Should be falsifiable against pillar thresholds.",
        },
        {
            "section": "Evidence posture",
            "metric": "Thesis pillars / evidence rows",
            "value": f"{count_rows('thesis_pillars')} / {count_rows('evidence_ledger')}",
            "notes": "Low evidence density should keep the tracker screen-grade.",
        },
        {
            "section": "Evidence posture",
            "metric": "Sources / open questions",
            "value": f"{count_rows('sources')} / {count_rows('open_questions')}",
            "notes": "Open questions stay visible until resolved.",
        },
        {
            "section": "Workbook map",
            "metric": "dashboard",
            "value": "Original input dashboard table",
            "notes": "Cover is the senior-review view; dashboard preserves source structure.",
        },
        {
            "section": "Workbook map",
            "metric": "thesis_pillars / evidence_ledger / kpi_tracker",
            "value": "Core proof stack",
            "notes": "Use to validate whether the thesis is improving or breaking.",
        },
        {
            "section": "Workbook map",
            "metric": "catalyst_calendar / estimate_revisions / model_changelog",
            "value": "Forward watchlist and model impact",
            "notes": "Use to keep the tracker live after events.",
        },
    ]


def write_xlsx(path: Path, section_rows: dict[str, tuple[list[str], list[dict[str, Any]]]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for --xlsx-out. CSV outputs were still created."
        ) from exc

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    cover_ws = wb.create_sheet("Cover")
    cover_fields = ["section", "metric", "value", "notes"]
    cover_ws.append(cover_fields)
    for cell in cover_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in tracker_cover_rows(section_rows):
        cover_ws.append([row.get(field, "") for field in cover_fields])
    cover_ws.freeze_panes = "A2"
    for col, width in {"A": 22, "B": 34, "C": 42, "D": 76}.items():
        cover_ws.column_dimensions[col].width = width
    for section, (fields, rows) in section_rows.items():
        ws = wb.create_sheet(section[:31])
        if rows and set(rows[0].keys()) == {"field", "value"}:
            fields = ["field", "value"]
        ws.append(fields)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            ws.append([row.get(field, "") for field in fields])
        ws.freeze_panes = "A2"
    wb.save(path)


def load_payload(path: Path | None) -> dict[str, Any]:
    if path is None:
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("input JSON must be an object")
    return data


def output_paths(output_dir: Path, xlsx_out: Path | None) -> dict[str, str]:
    paths = {section: str(output_dir / f"{section}.csv") for section in SECTIONS}
    paths["run_log"] = str(output_dir / "run_log.json")
    paths["manifest"] = str(output_dir / "manifest.json")
    if xlsx_out:
        paths["xlsx"] = str(xlsx_out)
    return paths


def output_manifest(
    paths: dict[str, str],
    *,
    required_xlsx: bool = False,
    primary_xlsx: bool = False,
) -> list[dict[str, Any]]:
    manifest = []
    for key, artifact_path in paths.items():
        is_primary = key == "xlsx" and primary_xlsx
        is_machine = key in {"run_log", "manifest"}
        manifest.append(
            {
                "key": key,
                "path": artifact_path,
                "required": key != "xlsx" or required_xlsx,
                "written": key in {"run_log", "manifest"} or Path(artifact_path).exists(),
                "description": "Thesis tracker artifact.",
                "artifact_role": "primary_human_deliverable"
                if is_primary
                else ("machine_support" if is_machine else "support_artifact"),
                "hidden_unless_requested": not is_primary,
            }
        )
    return manifest


def source_basis_from_payload(payload: dict[str, Any]) -> list[dict[str, Any]]:
    rows = rows_for_section(payload, "sources", SECTIONS["sources"])
    return [
        {
            "source_id": row.get("source_id", ""),
            "source_name": row.get("source_name", ""),
            "source_type": row.get("source_type", ""),
            "as_of_date": row.get("date_as_of", ""),
            "reliability": row.get("reliability", ""),
            "used_for": row.get("used_for", ""),
        }
        for row in rows
        if row.get("source_id") or row.get("source_name")
    ]


def write_run_log(output_dir: Path, run_log: dict[str, Any]) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "run_log.json").write_text(json.dumps(run_log, indent=2) + "\n", encoding="utf-8")
    primary = run_log.get("primary_human_deliverable")
    manifest = {
        "outputs": run_log["output_manifest"],
        "primary_human_deliverable": primary,
        "human_deliverables": [primary] if primary else [],
        "support_artifacts": [
            row["path"]
            for row in run_log["output_manifest"]
            if row.get("artifact_role") != "primary_human_deliverable"
        ],
        "support_artifacts_user_visible_default": False,
        "final_response_guidance": {
            "lead_with": "primary_human_deliverable" if primary else "concise_status",
            "mention_support_artifacts": "only_briefly_unless_requested",
        },
    }
    (output_dir / "manifest.json").write_text(
        json.dumps(manifest, indent=2) + "\n", encoding="utf-8"
    )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Materialize a public-equity-investing thesis tracker CSV bundle."
    )
    parser.add_argument("input_json", nargs="?", type=Path)
    parser.add_argument("--output-dir", type=Path, default=Path("output"))
    parser.add_argument("--xlsx-out", type=Path)
    args = parser.parse_args()

    args.output_dir.mkdir(parents=True, exist_ok=True)
    paths = output_paths(args.output_dir, args.xlsx_out)
    try:
        payload = load_payload(args.input_json)
    except Exception as exc:
        run_log = {
            "status": "failed",
            "model_status": "not-decision-ready",
            "artifact_level": "csv_tracker_bundle",
            "workbook_mode": "csv_tracker_bundle",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "input_json": str(args.input_json) if args.input_json else None,
            "source_basis": [],
            "hard_failures": [str(exc)],
            "warnings": [],
            "outputs": paths,
            "output_manifest": output_manifest(paths, required_xlsx=bool(args.xlsx_out)),
        }
        write_run_log(args.output_dir, run_log)
        print(f"ERROR: {exc}")
        return 1

    section_rows: dict[str, tuple[list[str], list[dict[str, Any]]]] = {}
    try:
        for section, fields in SECTIONS.items():
            rows = rows_for_section(payload, section, fields)
            write_csv(args.output_dir / f"{section}.csv", rows, fields)
            section_rows[section] = (fields, rows)
    except Exception as exc:
        run_log = {
            "status": "failed",
            "model_status": "not-decision-ready",
            "artifact_level": "csv_tracker_bundle",
            "workbook_mode": "csv_tracker_bundle",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "input_json": str(args.input_json) if args.input_json else None,
            "source_basis": source_basis_from_payload(payload),
            "hard_failures": [f"could not write output: {exc}"],
            "warnings": [],
            "outputs": paths,
            "output_manifest": output_manifest(paths, required_xlsx=bool(args.xlsx_out)),
        }
        write_run_log(args.output_dir, run_log)
        print(f"ERROR: could not write output: {exc}")
        return 1

    run_log = {
        "status": "completed",
        "model_status": "senior-review-ready",
        "artifact_level": "csv_tracker_bundle",
        "workbook_mode": "csv_tracker_bundle",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "artifact_mode": "csv_tracker_bundle",
        "input_json": str(args.input_json) if args.input_json else None,
        "sections": sorted(SECTIONS),
        "source_basis": source_basis_from_payload(payload),
        "hard_failures": [],
        "xlsx_requested": bool(args.xlsx_out),
        "xlsx_written": False,
        "outputs": paths,
        "warnings": [],
    }
    if args.xlsx_out:
        try:
            write_xlsx(args.xlsx_out, section_rows)
            run_log["xlsx_written"] = True
            run_log["xlsx_path"] = str(args.xlsx_out)
        except Exception as exc:
            run_log["status"] = "failed"
            run_log["model_status"] = "not-decision-ready"
            run_log["hard_failures"].append(str(exc))
            run_log["warnings"].append(str(exc))
            run_log["output_manifest"] = output_manifest(paths, required_xlsx=True)
            write_run_log(args.output_dir, run_log)
            print(f"ERROR: {exc}")
            return 1

    if run_log["xlsx_written"]:
        run_log["primary_human_deliverable"] = str(args.xlsx_out)
        run_log["artifact_mode"] = "workbook"
        run_log["workbook_mode"] = "xlsx_tracker_workbook"
    else:
        run_log["primary_human_deliverable"] = None

    run_log["output_manifest"] = output_manifest(
        paths,
        required_xlsx=bool(args.xlsx_out),
        primary_xlsx=bool(run_log["xlsx_written"]),
    )
    write_run_log(args.output_dir, run_log)
    print(f"Wrote thesis tracker CSV bundle to {args.output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
