"""Workbook naming and formatting helpers for excel-data-cleaner."""

from __future__ import annotations

import re
from pathlib import Path


def safe_excel_sheet_name(base: str, used: set[str]) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "_", base).strip() or "sheet"
    name = name[:31]
    original = name
    i = 2
    while name in used:
        suffix = f"_{i}"
        name = (original[: 31 - len(suffix)] + suffix)[:31]
        i += 1
    used.add(name)
    return name


def autosize_and_table(path: Path, sheet_formats: dict[str, dict[str, str]]) -> None:
    """Apply light Excel presentation polish after pandas writes the workbook."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row >= 1 and max_col >= 1:
            for cell in ws[1]:
                cell.style = "Headline 4"
            if headers_are_table_ready(ws, max_row, max_col):
                add_table(ws, max_col, max_row)
        apply_column_formats(ws, sheet_formats.get(ws.title, {}), max_row, max_col)
        autosize_columns(ws, max_row, max_col, get_column_letter)
    wb.save(path)


def headers_are_table_ready(ws, max_row: int, max_col: int) -> bool:
    return max_row >= 2 and all(
        ws.cell(1, col_idx).value not in (None, "") for col_idx in range(1, max_col + 1)
    )


def add_table(ws, max_col: int, max_row: int) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table_name = re.sub(r"[^A-Za-z0-9_]", "_", f"tbl_{ws.title}")[:240]
    if not table_name or table_name[0].isdigit():
        table_name = f"tbl_{table_name}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    try:
        ws.add_table(table)
    except ValueError:
        pass


def apply_column_formats(
    ws,
    formats: dict[str, str],
    max_row: int,
    max_col: int,
) -> None:
    header_to_col = {str(ws.cell(1, col_idx).value): col_idx for col_idx in range(1, max_col + 1)}
    for header, fmt in formats.items():
        col_idx = header_to_col.get(header)
        if col_idx:
            for row_idx in range(2, max_row + 1):
                ws.cell(row_idx, col_idx).number_format = fmt


def autosize_columns(ws, max_row: int, max_col: int, get_column_letter) -> None:
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 10
        for row_idx in range(1, min(max_row, 200) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 60))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 64)
