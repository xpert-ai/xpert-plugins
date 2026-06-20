#!/usr/bin/env python3
"""Conservative first-pass cleaner for messy CSV/XLSX data.

Creates an audit-ready workbook with cleaned data, raw source, data dictionary,
quality checks, and assumptions/audit sheets. This is a helper for the skill's
analyst workflow, not a replacement for human/AI review of ambiguous business rules.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

SCRIPT_DIR = Path(__file__).resolve().parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from workbook_formatting import autosize_and_table, safe_excel_sheet_name

BLANK_TOKENS = {
    "",
    "-",
    "--",
    "n/a",
    "na",
    "null",
    "none",
    "not available",
    "#n/a",
    "nan",
}
TOTAL_RE = re.compile(r"\b(grand\s+total|subtotal|sub-total|total)\b", re.I)
ID_HINT_RE = re.compile(
    r"\b(id|code|sku|zip|postal|phone|cusip|isin|sedol|ticker|account\s*#|invoice\s*#|po\s*#|employee)\b",
    re.I,
)
CURRENCY_RE = re.compile(r"[$€£¥]|\b(usd|eur|gbp|jpy|cad|aud|chf|cny)\b", re.I)
PERCENT_RE = re.compile(r"%|\bbps\b|\bbasis\s+points\b", re.I)
NUMERIC_RE = re.compile(
    r"^\(?[+-]?\s*[$€£¥]?\s*\d{1,3}(?:,\d{3})*(?:\.\d+)?\s*\)?$|^\(?[+-]?\s*[$€£¥]?\s*\d+(?:\.\d+)?\s*\)?$"
)

DOMAIN_REQUIRED_HINTS: dict[str, list[str]] = {
    "financial_statement_reporting": [
        "revenue",
        "ebitda",
        "eps",
        "fiscal",
        "segment",
        "currency",
        "period",
    ],
    "investing_markets": [
        "ticker",
        "security",
        "cusip",
        "isin",
        "date",
        "price",
        "return",
        "currency",
    ],
    "credit_markets_handoff": [
        "issuer",
        "debt",
        "maturity",
        "coupon",
        "yield",
        "spread",
        "rating",
        "covenant",
    ],
    "portfolio_risk": [
        "portfolio",
        "position",
        "weight",
        "exposure",
        "p&l",
        "benchmark",
        "factor",
    ],
    "consensus_provider_export": [
        "consensus",
        "estimate",
        "provider",
        "actual",
        "forecast",
        "revision",
    ],
    "event_driven": [
        "deal",
        "consideration",
        "spread",
        "probability",
        "closing",
        "regulatory",
        "break",
    ],
}


def ensure_dependencies() -> None:
    """Load runtime-only dependencies after argparse handles --help."""
    global pd
    try:
        import openpyxl  # noqa: F401
        import pandas as pandas
    except ImportError as exc:
        print(
            "ERROR: pandas and openpyxl are required to clean tabular data.",
            file=sys.stderr,
        )
        print(
            "Install with: python -m pip install -r scripts/requirements.txt",
            file=sys.stderr,
        )
        raise SystemExit(2) from exc
    pd = pandas


def norm_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ").strip())


def add_cover_sheet(
    path: Path,
    *,
    input_path: Path,
    summary_rows: list[dict[str, Any]],
    quality_checks: list[dict[str, Any]],
    audit_rows: list[dict[str, Any]],
    dedupe_policy: str,
    header_style: str,
    remove_subtotals: bool,
) -> None:
    """Insert a first-tab dashboard cover after pandas materializes the workbook."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    wb = load_workbook(path)
    if "Cover" in wb.sheetnames:
        del wb["Cover"]
    ws = wb.create_sheet("Cover", 0)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    rows = [
        {
            "section": "Header",
            "metric": "Input file",
            "value": str(input_path),
            "notes": "Raw source tabs are preserved behind this cover.",
        },
        {
            "section": "Status",
            "metric": "Workbook mode",
            "value": "cleaned_data_export",
            "notes": "Deterministic first-pass cleaner; reviewer judgment is still required.",
        },
        {
            "section": "Dataset overview",
            "metric": "Source sheets cleaned",
            "value": len(summary_rows),
            "notes": "See summary for per-sheet row/column counts.",
        },
        {
            "section": "Dataset overview",
            "metric": "Clean rows / columns",
            "value": f"{sum(int(r.get('clean_rows', 0) or 0) for r in summary_rows)} / {sum(int(r.get('clean_columns', 0) or 0) for r in summary_rows)}",
            "notes": "Aggregate across all cleaned sheets.",
        },
        {
            "section": "Quality dashboard",
            "metric": "Quality issues",
            "value": len(quality_checks),
            "notes": "See quality_checks before using the workbook for modeling.",
        },
        {
            "section": "Quality dashboard",
            "metric": "Fatal / warning issues",
            "value": sum(
                1
                for item in quality_checks
                if str(item.get("severity", "")).lower() in {"fatal", "warning"}
            ),
            "notes": "Fatal issues should block downstream use.",
        },
        {
            "section": "Cleaning actions",
            "metric": "Audit log rows",
            "value": len(audit_rows),
            "notes": "Every inferred transformation should be visible in assumptions_audit.",
        },
        {
            "section": "Cleaning actions",
            "metric": "Dedupe / header / subtotal policy",
            "value": f"{dedupe_policy} / {header_style} / {'remove' if remove_subtotals else 'preserve'}",
            "notes": "Confirm this matches the intended downstream use.",
        },
        {
            "section": "Source posture",
            "metric": "Raw source preservation",
            "value": "Preserved",
            "notes": "Raw sheets should remain unmodified except readability formatting.",
        },
        {
            "section": "Workbook map",
            "metric": "clean_data sheets",
            "value": "Analysis-ready output",
            "notes": "Use only after reviewing quality_checks and assumptions_audit.",
        },
        {
            "section": "Workbook map",
            "metric": "raw_source sheets",
            "value": "Original inputs",
            "notes": "Use for audit trail and tie-out.",
        },
        {
            "section": "Workbook map",
            "metric": "summary / data_dictionary / quality_checks / assumptions_audit",
            "value": "Reviewer dashboard and audit trail",
            "notes": "These tabs keep source gaps and cleaning assumptions visible.",
        },
    ]
    ws.append(["section", "metric", "value", "notes"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows:
        ws.append([row["section"], row["metric"], row["value"], row["notes"]])
    ws.freeze_panes = "A2"
    for col, width in {"A": 22, "B": 34, "C": 44, "D": 82}.items():
        ws.column_dimensions[col].width = width
    wb.save(path)


def is_blank(value: Any) -> bool:
    return norm_cell(value).lower() in BLANK_TOKENS


def clean_text_value(value: Any) -> Any:
    text = norm_cell(value)
    if text.lower() in BLANK_TOKENS:
        return None
    return text


def map_dataframe(df: pd.DataFrame, func):
    mapper = getattr(df, "map", None)
    if mapper is not None:
        return mapper(func)
    return df.applymap(func)


def snake_case(text: str) -> str:
    text = norm_cell(text).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "column"


def display_header(text: str) -> str:
    text = norm_cell(text)
    text = re.sub(r"[_\-]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return "Column"
    keep_upper = {
        "id",
        "sku",
        "arr",
        "mrr",
        "acv",
        "tcv",
        "cogs",
        "ebitda",
        "api",
        "sla",
        "po",
        "gl",
        "fy",
    }
    words = []
    for word in text.split(" "):
        stripped = re.sub(r"[^A-Za-z0-9]", "", word).lower()
        if stripped in keep_upper:
            words.append(stripped.upper())
        else:
            words.append(word[:1].upper() + word[1:].lower() if word else word)
    return " ".join(words)


def unique_headers(
    raw_headers: Iterable[Any], style: str
) -> tuple[list[str], list[dict[str, Any]]]:
    seen: dict[str, int] = defaultdict(int)
    headers: list[str] = []
    notes: list[dict[str, Any]] = []
    for idx, raw in enumerate(raw_headers, start=1):
        original = norm_cell(raw)
        if not original:
            original = f"column_{idx}"
            notes.append({"issue": "blank_header", "column_index": idx, "cleaned": original})
        base = snake_case(original) if style == "snake" else display_header(original)
        key = base.lower()
        seen[key] += 1
        cleaned = (
            base
            if seen[key] == 1
            else f"{base}_{seen[key]}"
            if style == "snake"
            else f"{base} {seen[key]}"
        )
        if seen[key] > 1:
            notes.append(
                {
                    "issue": "duplicate_header",
                    "column_index": idx,
                    "original": original,
                    "cleaned": cleaned,
                }
            )
        headers.append(cleaned)
    return headers, notes


def cell_has_letters(text: str) -> bool:
    return bool(re.search(r"[A-Za-z]", text))


def looks_numeric(text: str) -> bool:
    if not text:
        return False
    return bool(NUMERIC_RE.match(text.replace("%", "")))


def header_score(df: pd.DataFrame, row_idx: int) -> float:
    row = [norm_cell(v) for v in df.iloc[row_idx].tolist()]
    nonempty = [v for v in row if v]
    if len(nonempty) < 2:
        return -1000.0
    unique_ratio = len({v.lower() for v in nonempty}) / max(len(nonempty), 1)
    text_ratio = sum(cell_has_letters(v) for v in nonempty) / max(len(nonempty), 1)
    numeric_ratio = sum(looks_numeric(v) for v in nonempty) / max(len(nonempty), 1)
    next_ratio = 0.0
    if row_idx + 1 < len(df):
        next_vals = [norm_cell(v) for v in df.iloc[row_idx + 1].tolist()]
        next_ratio = sum(bool(v) for v in next_vals) / max(len(next_vals), 1)
    total_penalty = 2.0 if any(TOTAL_RE.search(v) for v in nonempty) else 0.0
    return (
        (len(nonempty) * 0.15)
        + (unique_ratio * 2.0)
        + (text_ratio * 3.0)
        + (next_ratio * 1.5)
        - (numeric_ratio * 2.0)
        - total_penalty
    )


def detect_header_row(df: pd.DataFrame) -> int:
    if df.empty:
        return 0
    scores = [(header_score(df, i), i) for i in range(min(len(df), 30))]
    scores.sort(reverse=True)
    return scores[0][1]


def read_input(path: Path, sheet: str | None = None) -> dict[str, pd.DataFrame]:
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xlsm", ".xls"}:
        data = pd.read_excel(
            path,
            sheet_name=sheet if sheet else None,
            header=None,
            dtype=object,
            keep_default_na=False,
        )
        if isinstance(data, pd.DataFrame):
            return {sheet or "sheet1": data}
        return {str(k): v for k, v in data.items()}
    if ext in {".csv", ".tsv", ".txt"}:
        sep = "\t" if ext == ".tsv" else None
        try:
            df = pd.read_csv(
                path,
                header=None,
                dtype=object,
                keep_default_na=False,
                sep=sep,
                engine="python",
                encoding="utf-8-sig",
            )
        except pd.errors.ParserError:
            df = pd.read_csv(
                path,
                header=None,
                dtype=object,
                keep_default_na=False,
                sep=sep,
                engine="python",
                encoding="utf-8-sig",
                on_bad_lines="warn",
            )
        return {"csv": df}
    raise ValueError(f"unsupported input type: {ext}")


def parse_number(value: Any) -> float | None:
    s = norm_cell(value)
    if not s or s.lower() in BLANK_TOKENS:
        return None
    s = CURRENCY_RE.sub("", s)
    s = s.replace(",", "").replace("%", "").strip()
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("() ")
    try:
        number = float(s)
        return -number if neg else number
    except ValueError:
        return None


def parse_percent(value: Any) -> float | None:
    s = norm_cell(value).lower()
    if not s or s in BLANK_TOKENS:
        return None
    if "bps" in s or "basis" in s:
        n = parse_number(s.replace("basis points", "").replace("bps", ""))
        return None if n is None else n / 10000.0
    n = parse_number(s)
    if n is None:
        return None
    if "%" in s:
        return n / 100.0
    # Leave ambiguous bare numbers as-is only if they already look like decimal percentages.
    return n if abs(n) <= 1 else None


def parse_bool(value: Any) -> bool | None:
    s = norm_cell(value).lower()
    if s in {"true", "yes", "y", "1"}:
        return True
    if s in {"false", "no", "n", "0"}:
        return False
    return None


def parse_dates(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce")


def infer_column_type(header: str, series: pd.Series) -> dict[str, Any]:
    values = [norm_cell(v) for v in series.tolist()]
    nonblank = [v for v in values if v.lower() not in BLANK_TOKENS]
    sample_n = max(min(len(nonblank), 1000), 1)
    if not nonblank:
        return {"type": "empty", "confidence": 1.0, "format": "general"}

    header_id_like = bool(ID_HINT_RE.search(header))
    leading_zero = any(re.match(r"^0\d+", v) for v in nonblank[:1000])
    bool_ratio = sum(parse_bool(v) is not None for v in nonblank[:1000]) / sample_n
    pct_ratio = sum(parse_percent(v) is not None for v in nonblank[:1000]) / sample_n
    num_ratio = sum(parse_number(v) is not None for v in nonblank[:1000]) / sample_n
    currency_ratio = sum(bool(CURRENCY_RE.search(v)) for v in nonblank[:1000]) / sample_n
    date_candidates = [v for v in nonblank[:1000] if not looks_numeric(v)]
    if date_candidates:
        date_ratio = parse_dates(pd.Series(date_candidates)).notna().sum() / len(date_candidates)
    else:
        date_ratio = 0.0

    if header_id_like or leading_zero:
        return {"type": "text_identifier", "confidence": 0.90, "format": "@"}
    if bool_ratio >= 0.90:
        return {"type": "boolean", "confidence": bool_ratio, "format": "general"}
    if pct_ratio >= 0.80:
        return {"type": "percent", "confidence": pct_ratio, "format": "0.0%"}
    if currency_ratio >= 0.40 and num_ratio >= 0.80:
        return {
            "type": "currency_amount",
            "confidence": min(0.95, (currency_ratio + num_ratio) / 2),
            "format": "#,##0.00",
        }
    if num_ratio >= 0.88:
        decimals = any(
            (parse_number(v) is not None and abs(parse_number(v) - round(parse_number(v))) > 1e-9)
            for v in nonblank[:1000]
        )
        return {
            "type": "number",
            "confidence": num_ratio,
            "format": "#,##0.00" if decimals else "#,##0",
        }
    if date_ratio >= 0.85:
        return {
            "type": "date_or_datetime",
            "confidence": date_ratio,
            "format": "yyyy-mm-dd",
        }
    if 0.30 <= num_ratio < 0.88 or 0.30 <= date_ratio < 0.85:
        return {
            "type": "mixed",
            "confidence": max(num_ratio, date_ratio),
            "format": "general",
        }
    return {"type": "text", "confidence": 0.70, "format": "general"}


def convert_column(series: pd.Series, col_type: str) -> pd.Series:
    cleaned = series.map(clean_text_value)
    if col_type in {"text", "text_identifier", "mixed", "empty"}:
        return cleaned
    if col_type == "boolean":
        return cleaned.map(lambda v: parse_bool(v) if v is not None else None)
    if col_type == "percent":
        return cleaned.map(lambda v: parse_percent(v) if v is not None else None)
    if col_type in {"number", "currency_amount"}:
        return cleaned.map(lambda v: parse_number(v) if v is not None else None)
    if col_type == "date_or_datetime":
        return parse_dates(cleaned)
    return cleaned


def likely_subtotal_mask(df: pd.DataFrame) -> pd.Series:
    if df.empty:
        return pd.Series([], dtype=bool)

    def row_has_total(row: pd.Series) -> bool:
        vals = [norm_cell(v) for v in row.tolist()[:5]]
        return any(TOTAL_RE.search(v) for v in vals if v)

    return df.apply(row_has_total, axis=1)


def infer_domain(sheet_name: str, headers: list[str], override: str) -> str:
    if override and override != "auto":
        return override
    text = " ".join([sheet_name] + headers).lower()
    scores: dict[str, int] = {}
    for domain, hints in DOMAIN_REQUIRED_HINTS.items():
        scores[domain] = sum(1 for h in hints if re.search(r"\b" + re.escape(h) + r"\b", text))
    best, score = max(scores.items(), key=lambda x: x[1])
    return best if score else "general"


def required_field_checks(
    domain: str, clean_df: pd.DataFrame, sheet_name: str
) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    if domain == "credit_markets_handoff":
        return [
            {
                "severity": "warning",
                "issue_type": "route_to_credit_markets",
                "sheet": sheet_name,
                "field": "domain",
                "affected_count": len(clean_df),
                "description": "Credit instrument table detected; route bonds/loans/CDS/covenants/recovery/debt-security analysis to Credit Markets unless this is only equity-risk context.",
                "recommended_action": "preserve source columns and hand off to Credit Markets for credit analysis; keep only equity-risk signal use in Public Equity Investing.",
            }
        ]
    hints = DOMAIN_REQUIRED_HINTS.get(domain, [])
    lower_cols = {c.lower(): c for c in clean_df.columns}
    for hint in hints:
        matches = [actual for lower, actual in lower_cols.items() if hint in lower]
        for col in matches[:2]:
            missing = int(clean_df[col].isna().sum())
            if missing:
                checks.append(
                    {
                        "severity": "warning",
                        "issue_type": "missing_domain_field_values",
                        "sheet": sheet_name,
                        "field": col,
                        "affected_count": missing,
                        "description": f"{missing} rows are missing values in a domain-relevant field.",
                        "recommended_action": "review whether these rows are incomplete or need enrichment from the source system.",
                    }
                )
    return checks


def apply_header_row(
    sheet_name: str,
    raw_df: pd.DataFrame,
    args: argparse.Namespace,
    audit: list[dict[str, Any]],
    checks: list[dict[str, Any]],
) -> pd.DataFrame:
    header_idx = args.header_row - 1 if args.header_row else detect_header_row(raw_df)
    raw_headers = raw_df.iloc[header_idx].tolist() if len(raw_df) else []
    headers, header_notes = unique_headers(raw_headers, args.header_style)
    data = raw_df.iloc[header_idx + 1 :].copy() if len(raw_df) else pd.DataFrame()
    if len(headers) == data.shape[1]:
        data.columns = headers
    audit.append(
        {
            "step": "header_detection",
            "action": f"used row {header_idx + 1} as header",
            "basis": "user-specified header row"
            if args.header_row
            else "automatic header inference",
            "affected_sheet": sheet_name,
            "affected_field": "all",
            "affected_rows_or_count": 1,
            "risk_level": "low" if args.header_row else "medium",
            "notes": "verify if the source has multi-row headers or report titles.",
        }
    )
    for note in header_notes:
        checks.append(
            {
                "severity": "warning",
                "issue_type": note["issue"],
                "sheet": sheet_name,
                "field": note.get("cleaned"),
                "affected_count": 1,
                "description": str(note),
                "recommended_action": "review header mapping in data_dictionary.",
            }
        )
    return data


def remove_empty_structure(
    sheet_name: str,
    data: pd.DataFrame,
    audit: list[dict[str, Any]],
) -> pd.DataFrame:
    before_rows, before_cols = data.shape
    data = data.dropna(how="all")
    data = data.loc[:, [not all(is_blank(v) for v in data[col].tolist()) for col in data.columns]]
    data = map_dataframe(data, clean_text_value)
    data = data.dropna(how="all")
    after_rows, after_cols = data.shape
    if before_rows != after_rows or before_cols != after_cols:
        audit.append(
            {
                "step": "structural_cleanup",
                "action": "removed fully empty rows/columns from clean output",
                "basis": "safe structural cleanup; raw_source remains preserved",
                "affected_sheet": sheet_name,
                "affected_field": "all",
                "affected_rows_or_count": f"rows {before_rows}->{after_rows}, columns {before_cols}->{after_cols}",
                "risk_level": "low",
                "notes": "only fully empty rows/columns were removed.",
            }
        )
    return data


def handle_subtotals(
    sheet_name: str,
    data: pd.DataFrame,
    args: argparse.Namespace,
    checks: list[dict[str, Any]],
    audit: list[dict[str, Any]],
) -> pd.DataFrame:
    subtotal_mask = likely_subtotal_mask(data)
    subtotal_count = int(subtotal_mask.sum())
    if subtotal_count:
        checks.append(
            {
                "severity": "warning",
                "issue_type": "possible_total_or_subtotal_rows",
                "sheet": sheet_name,
                "field": "row",
                "affected_count": subtotal_count,
                "description": "rows contain labels such as total or subtotal.",
                "recommended_action": "remove from detail data only if they are report subtotal rows, not real records.",
            }
        )
        if args.remove_subtotals:
            data = data.loc[~subtotal_mask].copy()
            audit.append(
                {
                    "step": "subtotal_handling",
                    "action": "removed possible total/subtotal rows",
                    "basis": "--remove-subtotals option",
                    "affected_sheet": sheet_name,
                    "affected_field": "row",
                    "affected_rows_or_count": subtotal_count,
                    "risk_level": "medium",
                    "notes": "verify these were not legitimate records.",
                }
            )
    return data


def infer_and_convert_columns(
    sheet_name: str,
    data: pd.DataFrame,
    checks: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    dictionary: list[dict[str, Any]] = []
    for original_col in list(data.columns):
        info = infer_column_type(original_col, data[original_col])
        converted = convert_column(data[original_col], info["type"])
        data[original_col] = converted
        non_null = data[original_col].dropna()
        examples = [str(v) for v in non_null.head(5).tolist()]
        dictionary.append(
            {
                "source_sheet": sheet_name,
                "original_field": original_col,
                "clean_field": original_col,
                "inferred_type": info["type"],
                "excel_format": info["format"],
                "null_count": int(data[original_col].isna().sum()),
                "unique_count": int(non_null.nunique(dropna=True)) if len(non_null) else 0,
                "example_values": ", ".join(examples),
                "cleaning_notes": f"confidence={round(float(info['confidence']), 3)}",
                "business_notes": "",
            }
        )
        if info["type"] == "mixed":
            checks.append(
                {
                    "severity": "warning",
                    "issue_type": "mixed_type_column",
                    "sheet": sheet_name,
                    "field": original_col,
                    "affected_count": int(len(non_null)),
                    "description": "column contains mixed values that were left mostly as text.",
                    "recommended_action": "review source values and provide a conversion rule if this field should be numeric/date.",
                }
            )
    return dictionary


def handle_duplicates(
    sheet_name: str,
    data: pd.DataFrame,
    args: argparse.Namespace,
    checks: list[dict[str, Any]],
    audit: list[dict[str, Any]],
) -> pd.DataFrame:
    if args.dedupe == "exact" and not data.empty:
        before = len(data)
        data = data.drop_duplicates(keep="first")
        removed = before - len(data)
        if removed:
            audit.append(
                {
                    "step": "duplicate_handling",
                    "action": "removed exact duplicate rows",
                    "basis": "--dedupe exact; all cleaned values matched",
                    "affected_sheet": sheet_name,
                    "affected_field": "all",
                    "affected_rows_or_count": removed,
                    "risk_level": "low",
                    "notes": "potential non-exact duplicates are not merged by this script.",
                }
            )
            checks.append(
                {
                    "severity": "info",
                    "issue_type": "exact_duplicate_rows_removed",
                    "sheet": sheet_name,
                    "field": "all",
                    "affected_count": removed,
                    "description": "exact duplicates were removed from clean output.",
                    "recommended_action": "review if source row identity matters.",
                }
            )
    return data


def clean_sheet(
    sheet_name: str, raw_df: pd.DataFrame, args: argparse.Namespace
) -> tuple[pd.DataFrame, list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], str]:
    audit: list[dict[str, Any]] = []
    checks: list[dict[str, Any]] = []

    data = apply_header_row(sheet_name, raw_df, args, audit, checks)
    data = remove_empty_structure(sheet_name, data, audit)
    domain = infer_domain(sheet_name, list(data.columns), args.domain)
    data = handle_subtotals(sheet_name, data, args, checks, audit)
    dictionary = infer_and_convert_columns(sheet_name, data, checks)
    data = handle_duplicates(sheet_name, data, args, checks, audit)
    checks.extend(required_field_checks(domain, data, sheet_name))
    return data.reset_index(drop=True), dictionary, checks, audit, domain


def write_support_manifest(
    output_path: Path, *, standalone: bool, quality_checks: list[dict[str, Any]]
) -> None:
    output_dir = output_path.resolve().parent
    run_log_path = output_dir / "run_log.json"
    manifest_path = output_dir / "manifest.json"
    primary = str(output_path.resolve()) if standalone else None
    output_manifest = [
        {
            "key": "cleaned_workbook",
            "path": str(output_path.resolve()),
            "required": True,
            "written": output_path.exists(),
            "artifact_role": "primary_human_deliverable" if standalone else "support_artifact",
            "hidden_unless_requested": not standalone,
            "description": "Cleaned workbook for Public Equity Investing table support.",
        },
        {
            "key": "run_log",
            "path": str(run_log_path),
            "required": True,
            "written": True,
            "artifact_role": "support_artifact",
            "hidden_unless_requested": True,
            "description": "Cleaner run log.",
        },
        {
            "key": "manifest",
            "path": str(manifest_path),
            "required": True,
            "written": True,
            "artifact_role": "support_artifact",
            "hidden_unless_requested": True,
            "description": "Cleaner manifest.",
        },
    ]
    run_log = {
        "status": "completed",
        "model_status": "screen-grade" if quality_checks else "senior-review-ready",
        "artifact_level": "standalone_support_request"
        if standalone
        else "embedded_support_artifact",
        "workbook_mode": "cleaned_workbook",
        "primary_human_deliverable": primary,
        "support_artifacts_user_visible_default": False,
        "warnings": [
            row.get("description", "")
            for row in quality_checks
            if row.get("severity") in {"warning", "medium", "high"}
        ],
        "hard_failures": [
            row.get("description", "")
            for row in quality_checks
            if row.get("severity") in {"fatal", "critical"}
        ],
        "output_manifest": output_manifest,
        "final_response_guidance": {
            "lead_with": "cleaned_workbook" if standalone else "owning_workflow_hero_artifact",
            "mention_support_artifacts": "only_briefly_unless_requested",
        },
    }
    run_log_path.write_text(json.dumps(run_log, indent=2) + "\n", encoding="utf-8")
    manifest_path.write_text(
        json.dumps(
            {
                "outputs": output_manifest,
                "primary_human_deliverable": primary,
                "support_artifacts_user_visible_default": False,
                "final_response_guidance": run_log["final_response_guidance"],
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Conservatively clean CSV/XLSX data into an audit-ready workbook."
    )
    parser.add_argument("input", help="input .xlsx, .xls, .csv, .tsv, or .txt file")
    parser.add_argument("--output", "-o", default="cleaned.xlsx", help="output .xlsx path")
    parser.add_argument("--sheet", help="optional Excel sheet name to clean")
    parser.add_argument(
        "--domain",
        default="auto",
        help="auto, financial_statement_reporting, investing_markets, credit_markets_handoff, portfolio_risk, consensus_provider_export, event_driven, or general",
    )
    parser.add_argument(
        "--dedupe",
        choices=["none", "exact"],
        default="exact",
        help="duplicate policy for clean output",
    )
    parser.add_argument("--header-row", type=int, help="1-based header row override")
    parser.add_argument(
        "--header-style",
        choices=["display", "snake"],
        default="display",
        help="cleaned header style",
    )
    parser.add_argument(
        "--remove-subtotals",
        action="store_true",
        help="remove rows that appear to be totals/subtotals",
    )
    parser.add_argument(
        "--standalone-cleaned-workbook",
        action="store_true",
        help="Mark the cleaned workbook as the explicit standalone cleanup deliverable.",
    )
    args = parser.parse_args()
    ensure_dependencies()

    input_path = Path(args.input)
    output_path = Path(args.output)
    tables = read_input(input_path, args.sheet)

    all_dictionary: list[dict[str, Any]] = []
    all_checks: list[dict[str, Any]] = []
    all_audit: list[dict[str, Any]] = []
    summary_rows: list[dict[str, Any]] = []
    sheet_formats: dict[str, dict[str, str]] = {}
    used_sheet_names: set[str] = set()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for source_name, raw_df in tables.items():
            clean_df, dictionary, checks, audit, domain = clean_sheet(source_name, raw_df, args)
            clean_sheet_name = safe_excel_sheet_name(
                "clean_data" if len(tables) == 1 else f"clean_{source_name}",
                used_sheet_names,
            )
            raw_sheet_name = safe_excel_sheet_name(
                "raw_source" if len(tables) == 1 else f"raw_{source_name}",
                used_sheet_names,
            )
            clean_df.to_excel(writer, sheet_name=clean_sheet_name, index=False)
            raw_df.to_excel(writer, sheet_name=raw_sheet_name, index=False, header=False)
            fmt_map = {
                row["clean_field"]: row["excel_format"]
                for row in dictionary
                if row.get("excel_format") not in {None, "general"}
            }
            sheet_formats[clean_sheet_name] = fmt_map
            all_dictionary.extend(dictionary)
            all_checks.extend(checks)
            all_audit.extend(audit)
            summary_rows.append(
                {
                    "source_sheet": source_name,
                    "clean_sheet": clean_sheet_name,
                    "raw_sheet": raw_sheet_name,
                    "inferred_domain": domain,
                    "clean_rows": len(clean_df),
                    "clean_columns": len(clean_df.columns),
                    "quality_issue_count": len(checks),
                }
            )

        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="summary", index=False)
        pd.DataFrame(all_dictionary or [{"note": "no columns profiled"}]).to_excel(
            writer, sheet_name="data_dictionary", index=False
        )
        pd.DataFrame(
            all_checks
            or [
                {
                    "severity": "info",
                    "issue_type": "no_issues_logged",
                    "description": "no quality issues were logged by the deterministic cleaner",
                }
            ]
        ).to_excel(writer, sheet_name="quality_checks", index=False)
        all_audit.append(
            {
                "step": "workbook_creation",
                "action": "created cleaned workbook",
                "basis": "deterministic first-pass cleaning script",
                "affected_sheet": "all",
                "affected_field": "all",
                "affected_rows_or_count": len(tables),
                "risk_level": "low",
                "notes": f"generated_at={datetime.now(timezone.utc).isoformat()}; input={input_path}",
            }
        )
        pd.DataFrame(all_audit).to_excel(writer, sheet_name="assumptions_audit", index=False)

    add_cover_sheet(
        output_path,
        input_path=input_path,
        summary_rows=summary_rows,
        quality_checks=all_checks,
        audit_rows=all_audit,
        dedupe_policy=args.dedupe,
        header_style=args.header_style,
        remove_subtotals=args.remove_subtotals,
    )
    autosize_and_table(output_path, sheet_formats)
    write_support_manifest(
        output_path, standalone=args.standalone_cleaned_workbook, quality_checks=all_checks
    )
    print(f"wrote cleaned workbook to {output_path}")


if __name__ == "__main__":
    main()
